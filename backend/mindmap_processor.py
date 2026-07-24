import os
import json
import logging
import uuid
import pandas as pd
import re
from collections import Counter
from fastapi import APIRouter, File, UploadFile, HTTPException, Form, Header
from fastapi.responses import FileResponse
from pydantic import BaseModel
import tempfile

# --- Path Fix ---
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

import analyzer
import extractor
get_genai_client = analyzer.get_genai_client
extract_pdf_content = extractor.extract_pdf_content
configure_response_format = analyzer.configure_response_format
robust_json_decode = analyzer.robust_json_decode

import asyncio
from fastapi import BackgroundTasks

logger = logging.getLogger(__name__)
router = APIRouter()

# 全域任務註冊表
task_registry = {}

class MindMapConfig(BaseModel):
    app_area_count: str = "3~5"
    tech1_count: str = "3~5"
    tech2_count: str = "3~7"
    tech3_count: str = "3~7"
    efficacy_count: str = "3~6"
    file_id: str = ""

# --- Pydantic Schemas for MindMap Structured Outputs ---

class CategoryMappingItem(BaseModel):
    original_category: str
    generalized_category: str

class CategoryGeneralization(BaseModel):
    mappings: list[CategoryMappingItem]

class LabelDefinitionItem(BaseModel):
    label_name: str
    definition: str

class LabelDefinitions(BaseModel):
    definitions: list[LabelDefinitionItem]

class TechTreeNode(BaseModel):
    技術1階: str
    技術2階: list[str]

class Stage1Taxonomy(BaseModel):
    summary_title: str
    應用領域: list[str]
    功效節點: list[str]
    技術樹: list[TechTreeNode]

class PatentStage1Mapping(BaseModel):
    專利公開公告號: str
    技術路徑: list[list[str]]
    應用領域: list[str]
    功效節點: list[str]

class Stage1MappingResponse(BaseModel):
    patents: list[PatentStage1Mapping]

class Stage2LabelResponse(BaseModel):
    技術3階類別: list[str]

class PatentStage2Mapping(BaseModel):
    專利公開公告號: str
    技術3階: list[str]

class Stage2MappingResponse(BaseModel):
    patents: list[PatentStage2Mapping]

# --- Pydantic Schemas for Preprocessing ---

class PatentPreprocessItem(BaseModel):
    專利公開公告號: str
    AI技術簡述: str
    技術特徵手段: str
    解決的技術問題或技術效益: str
    初篩結果: str

class PreprocessBatchResponse(BaseModel):
    patents: list[PatentPreprocessItem]

class StartFromPreprocessRequest(BaseModel):
    file_id: str
    config: MindMapConfig

class AIAssistCriteriaRequest(BaseModel):
    criteria_type: str
    user_input: str

async def safe_query_gemini_with_backoff(prompt, provider, client, response_schema=None, max_retries=5):
    """
    帶有指數退避重試的 API 呼叫包裝，支援 JSON Schema 約束。
    使用 asyncio.to_thread 執行同步 SDK 呼叫以避免阻塞事件循環。
    """
    prompt_preview = str(prompt)[:80].replace('\n', ' ')
    for attempt in range(max_retries):
        try:
            config_params = configure_response_format(response_schema, provider)
            if provider == "openrouter":
                logger.info(f"[AI API] Calling OpenRouter (attempt {attempt+1}/{max_retries}) | prompt: {prompt_preview}...")
                def call_api():
                    payload = {
                        "model": "google/gemini-2.5-flash",
                        "messages": [{"role": "user", "content": prompt}],
                        "temperature": 0.1, "max_tokens": 8000,
                        **config_params
                    }
                    return analyzer.send_openrouter_request(payload, timeout=300.0)
                resp_data = await asyncio.to_thread(call_api)
                result = resp_data["choices"][0]["message"]["content"]
                logger.info(f"[AI API] OpenRouter responded successfully ({len(result)} chars).")
                return result
            else:
                logger.info(f"[AI API] Calling Gemini 2.5 Flash (attempt {attempt+1}/{max_retries}) | prompt: {prompt_preview}...")
                def call_api():
                    config_dict = {
                        "max_output_tokens": 8000,
                        "temperature": 0.1,
                        **config_params
                    }
                    return client.models.generate_content(
                        model="gemini-2.5-flash", contents=prompt,
                        config=config_dict
                    )
                resp = await asyncio.to_thread(call_api)
                result = resp.text
                logger.info(f"[AI API] Gemini responded successfully ({len(result)} chars).")
                return result
        except Exception as e:
            wait_time = (2 ** attempt) + 0.5
            logger.warning(f"API call failed (attempt {attempt + 1}/{max_retries}): {e}. Retrying in {wait_time}s...")
            if attempt == max_retries - 1:
                raise e
            await asyncio.sleep(wait_time)

def extract_patents_from_excel(file_path):
    df = pd.read_excel(file_path)
    return df.fillna("")

def df_to_ai_content(df):
    relevant_keywords = ["號", "名稱", "領域", "技術", "功效"]
    relevant_cols = [c for c in df.columns if any(k in str(c) for k in relevant_keywords)]
    ai_patents = []
    for _, row in df.iterrows():
        p_data = {str(c): str(row[c]) for c in relevant_cols if row[c] != ""}
        ai_patents.append(p_data)
    return json.dumps(ai_patents, ensure_ascii=False)

def refine_categories_with_ai(parent_name, subcategories, max_count, provider, client=None):
    """
    使用 AI 進行超限分類的「層級廣度自癒壓縮（AI-driven generalization）」。
    """
    subcats_str = ", ".join([f'"{s}"' for s in subcategories])
    prompt = f"""你是專利分類專家。在專利技術層級分類中，父類別「{parent_name}」下目前有以下 {len(subcategories)} 個子類別：
[{subcats_str}]

因為限制，子類別數量最多只能有 {max_count} 個。
請依據這些子類別的名稱與語意特徵，重新進行概念向上歸納（Generalization / 分類廣度調整），將它們歸併為最多 {max_count} 個標籤命名更寬廣、更具代表性的新子類別。

請輸出一個 JSON 物件，其中包含 key "mappings" (陣列)，陣列中每個元素為包含 "original_category" 與 "generalized_category" 的物件，將原來的子類別對應到歸納後的新子類別，例如：
{{
  "mappings": [
    {{"original_category": "原類別A", "generalized_category": "新歸納寬類別1"}},
    {{"original_category": "原類別B", "generalized_category": "新歸納寬類別1"}},
    {{"original_category": "原類別C", "generalized_category": "新歸納寬類別2"}}
  ]
}}
注意：對照表中 "generalized_category" 的相異值（新類別）數量不能超過 {max_count} 個。
ONLY output raw JSON. Do not include markdown blocks like ```json in the final response.
"""
    try:
        config_params = configure_response_format(CategoryGeneralization, provider)
        if provider == "openrouter":
            payload = {
                "model": "google/gemini-2.5-flash",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
                **config_params
            }
            resp_data = analyzer.send_openrouter_request(payload, timeout=60.0)
            response_text = resp_data["choices"][0]["message"]["content"]
        else:
            if client is None:
                client = get_genai_client()
            resp = client.models.generate_content(
                model="gemini-2.5-flash", contents=prompt,
                config={"temperature": 0.1, "response_mime_type": "application/json"}
            )
            response_text = resp.text

        mapping = json.loads(response_text)
        if isinstance(mapping, dict):
            return mapping
        return {}
    except Exception as e:
        logger.error(f"AI category refinement failed: {e}. Falling back to standard mapping.")
        return {}

def enforce_category_limits_hierarchical(patents: list, config: MindMapConfig, provider, client=None) -> list:
    """
    層階式限額強制器：
    1. 應用領域：限制在 app_area_count 內。
    2. 技術1階：限制在 tech1_count 內。
    3. 技術2階：對於每個技術1階，限制其擁有的技術2階子節點在 tech2_count 內。
    4. 技術3階：對於每個 (技術1階, 技術2階) 父節點對，限制其技術3階在 tech3_count 內。
    5. 功效節點：限制在 efficacy_count 內。
    
    使用 AI-driven semantic generalization 作為主要手段，若失敗則以機械合併作 fallback。
    """
    def safe_int(val, default):
        try:
            val_str = str(val).strip()
            if '~' in val_str:
                return int(val_str.split('~')[-1].strip())
            return int(val_str)
        except:
            return default

    lim_app = safe_int(config.app_area_count, 7)
    lim_t1 = safe_int(config.tech1_count, 5)
    lim_t2 = safe_int(config.tech2_count, 7)
    lim_t3 = safe_int(config.tech3_count, 5)
    lim_eff = safe_int(config.efficacy_count, 5)

    def apply_fallback_merge(items_set, frequencies_counter, limit, current_mapping):
        mapped_vals = set(current_mapping.values())
        if not current_mapping or len(mapped_vals) > limit:
            sorted_items = sorted(list(items_set), key=lambda x: frequencies_counter.get(x, 0), reverse=True)
            keep = set(sorted_items[:limit])
            fallback = sorted_items[0] if sorted_items else "其他"
            for v in items_set:
                if v not in keep and v not in current_mapping:
                    current_mapping[v] = fallback
        return current_mapping

    # Ensure all patents have basic list fields
    for p in patents:
        if "應用領域" not in p or not p["應用領域"]: p["應用領域"] = ["其他"]
        if "功效節點" not in p or not p["功效節點"]: p["功效節點"] = ["其他"]
        if "技術路徑" not in p:
            t1 = p.get("技術1階", ["其他"])
            t2 = p.get("技術2階", ["其他"])
            t3 = p.get("技術3階", ["其他"])
            t1_str = t1[0] if isinstance(t1, list) and t1 else str(t1)
            t2_str = t2[0] if isinstance(t2, list) and t2 else str(t2)
            t3_str = t3[0] if isinstance(t3, list) and t3 else str(t3)
            p["技術路徑"] = [[t1_str, t2_str, t3_str]]

    # 1. 應用領域 (Global limit)
    app_freq = Counter()
    for p in patents:
        for val in p.get("應用領域", []):
            app_freq[val] += 1
    if len(app_freq) > lim_app:
        mapping = refine_categories_with_ai("應用領域", list(app_freq.keys()), lim_app, provider, client)
        mapping = apply_fallback_merge(list(app_freq.keys()), app_freq, lim_app, mapping)
        for p in patents:
            p["應用領域"] = list(set([mapping.get(v, v) for v in p.get("應用領域", [])]))

    # 2. 功效節點 (Global limit)
    eff_freq = Counter()
    for p in patents:
        for val in p.get("功效節點", []):
            eff_freq[val] += 1
    if len(eff_freq) > lim_eff:
        mapping = refine_categories_with_ai("功效節點", list(eff_freq.keys()), lim_eff, provider, client)
        mapping = apply_fallback_merge(list(eff_freq.keys()), eff_freq, lim_eff, mapping)
        for p in patents:
            p["功效節點"] = list(set([mapping.get(v, v) for v in p.get("功效節點", [])]))

    # 3. 技術1階 (Global limit for level 1)
    t1_freq = Counter()
    for p in patents:
        for path in p.get("技術路徑", []):
            if len(path) > 0:
                t1_freq[path[0]] += 1
    if len(t1_freq) > lim_t1:
        mapping = refine_categories_with_ai("技術1階", list(t1_freq.keys()), lim_t1, provider, client)
        mapping = apply_fallback_merge(list(t1_freq.keys()), t1_freq, lim_t1, mapping)
        for p in patents:
            new_paths = []
            for path in p.get("技術路徑", []):
                if len(path) > 0:
                    path[0] = mapping.get(path[0], path[0])
                new_paths.append(path)
               # 5. 技術3階 (Per Technical 2 node)
    t2_to_t3 = {}
    t3_freq_per_t2 = {}
    for p in patents:
        for path in p.get("技術路徑", []):
            if len(path) > 2:
                t1, t2, t3 = path[0], path[1], path[2]
                t2_to_t3.setdefault((t1, t2), set()).add(t3)
                t3_freq_per_t2.setdefault((t1, t2), Counter())[t3] += 1

def query_definitions_for_taxonomy(taxonomy: dict, provider: str, client=None) -> dict:
    """
    Second-pass API call: generate ~60-char Traditional Chinese definitions for each taxonomy label.
    Separated from Stage 1 to avoid JSON truncation caused by increased output token length.
    Returns a flat dict {label_name: definition_text}. Non-critical: returns {} on failure.
    """
    import dotenv
    dotenv.load_dotenv(override=True)

    # Collect all labels from the taxonomy
    labels = []
    labels.extend(taxonomy.get("應用領域", []))
    labels.extend(taxonomy.get("功效節點", []))
    for t1_item in taxonomy.get("技術樹", []):
        t1_name = t1_item.get("技術1階", "")
        if t1_name:
            labels.append(t1_name)
        labels.extend([t2 for t2 in t1_item.get("技術2階", []) if t2])

    if not labels:
        return {}

    labels_str = "\n".join([f"- {label}" for label in labels])
    prompt = f"""你是專利分類專家。以下是一組專利分類標籤，請為每個標籤撰寫約 60 字的繁體中文定義說明。

分類標籤清單：
{labels_str}

【格式要求】：
- 輸出一個 JSON 物件，包含唯一的 key "definitions" (陣列)。
- 陣列中每個元素為包含 "label_name" 與 "definition" 的物件。
- 範例格式：
{{
  "definitions": [
    {{"label_name": "標籤名稱A", "definition": "此類別涵蓋相關技術，主要應用於特定場景，具有代表性特性。"}},
    {{"label_name": "標籤名稱B", "definition": "..."}}
  ]
}}
"""
    try:
        config_params = configure_response_format(LabelDefinitions, provider)
        response_text = ""
        if provider == "openrouter":
            payload = {
                "model": "google/gemini-2.5-flash",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
                "max_tokens": 16000,
                **config_params
            }
            resp_data = analyzer.send_openrouter_request(payload, timeout=300.0)
            response_text = resp_data["choices"][0]["message"]["content"]
        else:
            if client is None:
                client = get_genai_client()
            config_dict = {
                "max_output_tokens": 16000,
                "temperature": 0.1,
                **config_params
            }
            resp = client.models.generate_content(
                model="gemini-2.5-flash", contents=prompt,
                config=config_dict
            )
            response_text = resp.text

        parsed = robust_json_decode(response_text)
        definitions_dict = {}
        if isinstance(parsed, dict) and "definitions" in parsed:
            for item in parsed["definitions"]:
                if isinstance(item, dict) and "label_name" in item and "definition" in item:
                    definitions_dict[item["label_name"]] = item["definition"]
        elif isinstance(parsed, dict):
            # Fallback legacy parsing
            for k, v in parsed.items():
                if isinstance(v, str) and k != "definitions":
                    definitions_dict[k] = v
        return definitions_dict
    except Exception as e:
        logger.warning(f"Definitions generation failed (non-critical, will be empty): {e}")
        return {}


def clean_label_name(label: str, parent_label: str = None) -> str:
    """
    1. 技術節點若其上階命名已有技術名稱，其下一階的命名不須再重複相同字眼。
    2. 命名的總字數需不超過 12 個字。
    """
    if not label:
        return label
    label = label.strip()

    if parent_label:
        parent_label = parent_label.strip()
        if parent_label and parent_label in label:
            cleaned = label.replace(parent_label, "")
            cleaned = cleaned.strip("-_/、 ")
            if cleaned:
                label = cleaned

    if len(label) > 12:
        label = label[:12]

    return label


def clean_stage1_taxonomy(taxonomy: dict) -> dict:
    """
    清理 Stage 1 生成之分類結構中的標籤命名。
    """
    if not isinstance(taxonomy, dict):
        return taxonomy

    # 1. 應用領域
    if "應用領域" in taxonomy and isinstance(taxonomy["應用領域"], list):
        taxonomy["應用領域"] = [clean_label_name(app) for app in taxonomy["應用領域"] if app]

    # 2. 功效節點
    if "功效節點" in taxonomy and isinstance(taxonomy["功效節點"], list):
        taxonomy["功效節點"] = [clean_label_name(eff) for eff in taxonomy["功效節點"] if eff]

    # 3. 技術樹
    if "技術樹" in taxonomy and isinstance(taxonomy["技術樹"], list):
        for t1_item in taxonomy["技術樹"]:
            if not isinstance(t1_item, dict):
                continue
            t1_name = t1_item.get("技術1階", "")
            if t1_name:
                t1_name_cleaned = clean_label_name(t1_name)
                t1_item["技術1階"] = t1_name_cleaned

                # 技術2階 (以技術1階名稱為 parent)
                if "技術2階" in t1_item and isinstance(t1_item["技術2階"], list):
                    t2_list_cleaned = []
                    for t2 in t1_item["技術2階"]:
                        if t2:
                            t2_cleaned = clean_label_name(t2, parent_label=t1_name_cleaned)
                            t2_list_cleaned.append(t2_cleaned)
                    t1_item["技術2階"] = t2_list_cleaned

    return taxonomy


def query_gemini_stage1(text_content: str, config: MindMapConfig, df=None):
    import dotenv
    dotenv.load_dotenv(override=True)
    provider = os.environ.get("API_PROVIDER", "gemini").lower().strip().replace('"', '').replace("'", "")
    client = get_genai_client() if provider != "openrouter" else None

    # 優先抓取 "Publication Number" 欄位，其次抓含「號」的欄位
    pub_col_name = (
        next((c for c in df.columns if "publication number" in str(c).lower()), None)
        or next((c for c in df.columns if "號" in str(c)), None)
    ) if df is not None else None
    
    col_app = next((c for c in df.columns if "應用領域" in str(c)), "")
    col_eff = next((c for c in df.columns if "功效節點" in str(c)), "")
    col_brief = next((c for c in df.columns if "AI技術簡述" in str(c)), "")
    col_means = next((c for c in df.columns if "技術特徵手段" in str(c)), "")
    col_effect = next((c for c in df.columns if "解決的技術問題或技術效益" in str(c)), "")

    patents_list = ""
    if df is not None:
        for _, row in df.iterrows():
            p_no = str(row.get(pub_col_name, "N/A")).strip()
            bits = []
            if col_app and str(row[col_app]).strip(): bits.append(f"領域:{str(row[col_app]).strip()}")
            if col_eff and str(row[col_eff]).strip(): bits.append(f"功效:{str(row[col_eff]).strip()}")
            
            tech_elements = []
            if col_brief and str(row[col_brief]).strip(): tech_elements.append(f"簡述:{str(row[col_brief]).strip()}")
            if col_means and str(row[col_means]).strip(): tech_elements.append(f"手段:{str(row[col_means]).strip()}")
            if col_effect and str(row[col_effect]).strip(): tech_elements.append(f"效益:{str(row[col_effect]).strip()}")
            if tech_elements:
                bits.append(f"技術特徵:{' '.join(tech_elements)}")
                
            patents_list += f"[{p_no}] {' | '.join(bits)}\n"

    # Stage 1: Generate global level 1 & 2 taxonomy, application areas, and efficacy nodes.
    prompt = f"""你是專利分類專家，請對這批專利進行語意分類與樹狀結構建模。
請注意：本階段【只生成全域的分類樹架構，不要對具體專利進行映射或分配，也不要生成技術3階】。

【分類限制要求】：
1. 應用領域：預設分類數量限制為 {config.app_area_count} 個。
2. 功效節點：預設分類數量限制為 {config.efficacy_count} 個。
3. 技術層級樹（只到2階）：
   - 技術1階：總共分類為 {config.tech1_count} 個主要技術類別。
   - 技術2階（依附於技術1階）：在每個「技術1階」下，分類為 {config.tech2_count} 個子技術類別。

【分類標籤命名規則】：
1. 命名總字數限制：所有分類標籤（包括應用領域、功效節點、技術1階、技術2階）之命名總字數不得超過 12 個字。
2. 避免重複上階名稱：下階（技術2階）之命名不須且不要再重複其對應上階（技術1階）已有的技術名稱。例如：若「技術1階」命名為「光子積體電路」，則其下屬之「技術2階」不應命名為「光子積體電路電路設計」，而應簡化為「電路設計」。

【任務與格式規定】：
- 請構建全域分類目錄。
- 產出一個 summary_title 概括這批專利的核心技術主題，必須是完整、有意義且不超過12字的名詞短語（例如：「光子積體電路技術與應用」、「半導體先進封裝技術」）。
- 務必調整分類命名的廣度 (Generalization/Specialization) 使得輸出的分類總數不違反上述限制。

【格式】ONLY output a JSON object with keys: summary_title (string), 應用領域 (array of strings), 功效節點 (array of strings), 技術樹 (array of objects).
技術樹結構：
"技術樹": [
  {{
    "技術1階": "技術1階名稱A",
    "技術2階": ["子技術1", "子技術2"]
  }}
]

待分析清單：
{patents_list}
"""
    try:
        config_params = configure_response_format(Stage1Taxonomy, provider)
        response_text = ""
        if provider == "openrouter":
            payload = {
                "model": "google/gemini-2.5-flash",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1, "max_tokens": 8000,
                **config_params
            }
            resp_data = analyzer.send_openrouter_request(payload, timeout=600.0)
            response_text = resp_data["choices"][0]["message"]["content"]
        else:
            config_dict = {
                "max_output_tokens": 8000,
                "temperature": 0.1,
                **config_params
            }
            resp = client.models.generate_content(
                model="gemini-2.5-flash", contents=prompt,
                config=config_dict
            )
            response_text = resp.text

        parsed = robust_json_decode(response_text)
        if parsed:
            # 執行後處理命名規則清理 (12字上限及去除技術重複字眼)
            parsed = clean_stage1_taxonomy(parsed)
            # Second-pass: generate definitions in a separate call to avoid Stage 1 JSON truncation
            logger.info("Stage 1 taxonomy generated. Running second-pass definitions generation...")
            definitions = query_definitions_for_taxonomy(parsed, provider, client)
            parsed["定義說明"] = definitions
            logger.info(f"Definitions generated for {len(definitions)} labels.")
        return parsed
    except Exception as e:
        logger.error(f"Stage 1 LLM failed: {e}")
        return None

async def preprocess_single_patent_with_retry(
    p_no: str,
    row,
    title_col,
    abstract_col,
    novelty_col,
    use_col,
    advantage_col,
    claim_col,
    enable_screening: bool,
    screening_criteria: str,
    provider: str,
    client
) -> dict:
    """
    對單件專利進行預處理，最多重試 5 次。
    使用指數退避（Exponential Backoff）防止連續觸發 Rate Limit。
    5 次全部失敗後才回傳預設錯誤哨兵値。
    """
    title = str(row.get(title_col, "")).strip() if title_col else "無"
    abstract = str(row.get(abstract_col, "")).strip() if abstract_col else "無"
    novelty = str(row.get(novelty_col, "")).strip() if novelty_col else "無"
    use = str(row.get(use_col, "")).strip() if use_col else "無"
    advantage = str(row.get(advantage_col, "")).strip() if advantage_col else "無"
    claim = str(row.get(claim_col, "")).strip() if claim_col else "無"

    prompt = f"""你是專利分析與檢索專家。請針對以下 1 件專利進行分析，執行以下兩個任務：
1. 撰寫各約 50-60 字的繁體中文「AI技術簡述」、「技術特徵手段」與「解決的技術問題或技術效益」。
2. 【落入範圍初篩】：判定該專利是否落入範圍，結果限為 "Y" 或 "N"。無準則時全部判為 "Y"。

【篩選準則】：
{screening_criteria if enable_screening else "無篩選準則，請初篩判定為 Y"}

【寫作規範】：
- 「AI技術簡述」：說明該專利核心技術主題，約 50-60 字繁體中文。
- 「技術特徵手段」：具體列出其實現的關鍵特徵，約 50-60 字繁體中文。
- 「解決的技術問題或技術效益」：描述其克服的技術缺陷或帶來的效能提升，約 50-60 字繁體中文。
- 「專利公開公告號」必須與輸入的完全一致。

專利公開公告號: {p_no}
標題: {title}
專利摘要Abstract: {abstract}
DWPI Novelty: {novelty}
DWPI Use: {use}
DWPI Advantage: {advantage}
First Claim (Original language): {claim}
"""

    fallback = {
        "AI技術簡述": "專利資料預處理失敗，重試已達 5 次上限，無法自動生成技術簡述。",
        "技術特徵手段": "專利資料預處理失敗，重試已達 5 次上限，無法自動生成技術特徵手段。",
        "解決的技術問題或技術效益": "專利資料預處理失敗，重試已達 5 次上限，無法自動生成技術效益說明。",
        "初篩結果": "Y" if not enable_screening else "N"
    }

    for attempt in range(5):
        try:
            response_text = await safe_query_gemini_with_backoff(
                prompt, provider, client, response_schema=PatentPreprocessItem
            )
            parsed = robust_json_decode(response_text)
            if isinstance(parsed, dict) and parsed.get("AI技術簡述"):
                logger.info(f"Single-patent retry succeeded for {p_no} on attempt {attempt + 1}.")
                return {
                    "AI技術簡述": parsed.get("AI技術簡述", ""),
                    "技術特徵手段": parsed.get("技術特徵手段", ""),
                    "解決的技術問題或技術效益": parsed.get("解決的技術問題或技術效益", ""),
                    "初篩結果": "Y" if not enable_screening else ("Y" if str(parsed.get("初篩結果", "N")).upper() == "Y" else "N")
                }
        except Exception as e:
            wait = 2 ** attempt
            logger.warning(f"Single-patent retry [{attempt + 1}/5] failed for {p_no}: {e}. Waiting {wait}s...")
            await asyncio.sleep(wait)

    logger.error(f"Single-patent retry exhausted all 5 attempts for {p_no}. Returning fallback.")
    return fallback


async def run_preprocess_task(
    task_id: str,
    file_id: str,
    file_path: str,
    filename: str,
    enable_screening: bool,
    screening_criteria: str
):
    try:
        import dotenv
        dotenv.load_dotenv(override=True)
        provider = os.environ.get("API_PROVIDER", "gemini").lower().strip().replace('"', '').replace("'", "")
        client = get_genai_client() if provider != "openrouter" else None

        df = extract_patents_from_excel(file_path)
        total_count = len(df)

        title_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["標題", "title", "名稱", "name"])), None)
        abstract_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["摘要", "abstract"])), None)
        novelty_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["novelty", "新穎性"])), None)
        use_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["dwpi use", "use", "用途"])), None)
        advantage_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["advantage", "優點", "功效"])), None)
        claim_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["claim", "申請專利範圍", "權利要求", "first claim"])), None)
        # 優先抓取 "Publication Number" 欄位，其次抓含「號」或 "number" 的欄位
        pub_col_name = (
            next((c for c in df.columns if "publication number" in str(c).lower()), None)
            or next((c for c in df.columns if "號" in str(c).lower() or "number" in str(c).lower()), None)
            or df.columns[0]
        )

        # 偵測3個預處理輸出欄位 (若 Excel 中已存在)
        col_brief  = next((c for c in df.columns if "AI技術簡述" in str(c)), None)
        col_means  = next((c for c in df.columns if "技術特徵手段" in str(c)), None)
        col_effect = next((c for c in df.columns if "解決的技術問題或技術效益" in str(c)), None)
        col_screen = next((c for c in df.columns if "初篩結果" in str(c)), None)

        def _get_existing_fields(row):
            """從既有欄位取得3個預處理欄位值，回傳 (brief, means, effect) 字串三元組。"""
            brief  = str(row.get(col_brief,  "")).strip() if col_brief  else ""
            means  = str(row.get(col_means,  "")).strip() if col_means  else ""
            effect = str(row.get(col_effect, "")).strip() if col_effect else ""
            return brief, means, effect

        def _get_existing_screen(row):
            """從既有欄位取得初篩結果，若欄位不存在或值為空則回傳空字串。"""
            if col_screen is None:
                return ""
            val = str(row.get(col_screen, "")).strip().upper()
            return val if val in ("Y", "N") else ""

        checkpoint_path = os.path.join(checkpoint_dir, f"{file_id}_preprocess_checkpoint.json")
        checkpoint_data = {
            "completed_patents": {}
        }

        if os.path.exists(checkpoint_path):
            try:
                with open(checkpoint_path, "r", encoding="utf-8") as f:
                    checkpoint_data = json.load(f)
            except Exception as e:
                logger.error(f"Failed to load preprocess checkpoint: {e}")

        completed_patents = checkpoint_data.get("completed_patents", {})

        # 將待處理列分為三類：
        #   ai_full_rows       — 任一預處理欄位為空 → 需要AI全量預處理
        #   ai_screen_only_rows — 3欄均有值，但初篩結果為空且啟用初篩 → 只需AI做Y/N初篩
        #   skipped_rows        — 3欄均有值且不需初篩判斷 → 直接沿用既有值
        ai_full_rows = []
        ai_screen_only_rows = []
        skipped_rows = []

        for idx, row in df.iterrows():
            p_no = str(row.get(pub_col_name, "")).strip()
            if not p_no:
                p_no = f"ROW_{idx}"
            if p_no in completed_patents:
                continue  # 已有 checkpoint 結果，跳過
            brief, means, effect = _get_existing_fields(row)
            has_all_fields = bool(brief and means and effect)
            if not has_all_fields:
                ai_full_rows.append((idx, p_no, row))
            else:
                existing_screen = _get_existing_screen(row)
                if enable_screening and not existing_screen:
                    # 三欄均有值，但初篩結果缺漏且啟用初篩 → 只需AI判斷Y/N
                    ai_screen_only_rows.append((idx, p_no, row))
                else:
                    # 完全不需AI：沿用既有值（初篩結果若有值用原值，否則預設Y）
                    skipped_rows.append((idx, p_no, row))

        # 預填 skipped 專利到 completed_patents（直接沿用既有欄位值）
        for idx, p_no, row in skipped_rows:
            brief, means, effect = _get_existing_fields(row)
            existing_screen = _get_existing_screen(row)
            completed_patents[p_no] = {
                "AI技術簡述": brief,
                "技術特徵手段": means,
                "解決的技術問題或技術效益": effect,
                "初篩結果": existing_screen if existing_screen else "Y"
            }
            logger.info(f"專利 {p_no}：3個預處理欄位均有值，跳過AI預處理，直接沿用既有資料。")

        # 進度分母：需要AI處理的專利數（全量 + 僅初篩）
        pending_rows = ai_full_rows  # 統一使用 pending_rows 變數供後續批次邏輯使用
        ai_total_count = len(ai_full_rows) + len(ai_screen_only_rows)
        task_registry[task_id]["total_count"] = ai_total_count if ai_total_count > 0 else 1
        task_registry[task_id]["completed_count"] = len(completed_patents) - len(skipped_rows)

        batch_size = 10
        semaphore = asyncio.Semaphore(5)

        async def process_batch(batch):
            async with semaphore:
                prompt = f"""你是專利分析與檢索專家。請針對以下 10 件專利的技術特徵與公開文字進行分析，並執行以下兩個任務：
1. 為每一件專利分別撰寫各約 50-60 字的繁體中文「AI技術簡述」、「技術特徵手段」與「解決的技術問題或技術效益」。
2. 【落入範圍初篩作業】：判定該專利是否落入特定的篩選準則範圍。初篩判定結果限為落入："Y" 或不落入："N"。
   如果沒有提供篩選準則，請全部判定為 "Y"。

【篩選準則】：
{screening_criteria if enable_screening else "無篩選準則，請全部初篩判定為 Y"}

【寫作規範】：
- 「AI技術簡述」：說明該專利核心技術主題，約 50-60 字繁體中文。
- 「技術特徵手段」：具體列出其實現的關鍵硬體/軟體/程序特徵手段，約 50-60 字繁體中文。
- 「解決的技術問題或技術效益」：描述該專利克服的技術缺陷或帶來的效能提升，約 50-60 字繁體中文。
- 每件專利的「專利公開公告號」必須與輸入的「專利公開公告號」完全一致。

待分析專利列表：
"""
                for idx, p_no, row in batch:
                    title = str(row.get(title_col, "")).strip() if title_col else "無"
                    abstract = str(row.get(abstract_col, "")).strip() if abstract_col else "無"
                    novelty = str(row.get(novelty_col, "")).strip() if novelty_col else "無"
                    use = str(row.get(use_col, "")).strip() if use_col else "無"
                    advantage = str(row.get(advantage_col, "")).strip() if advantage_col else "無"
                    claim = str(row.get(claim_col, "")).strip() if claim_col else "無"

                    prompt += f"""
---
專利公開公告號: {p_no}
標題: {title}
專利摘要Abstract: {abstract}
DWPI Novelty: {novelty}
DWPI Use: {use}
DWPI Advantage: {advantage}
First Claim (Original language): {claim}
"""
                try:
                    response_text = await safe_query_gemini_with_backoff(prompt, provider, client, response_schema=PreprocessBatchResponse)
                    parsed = robust_json_decode(response_text)
                    batch_mapped = parsed.get("patents", [])
                    
                    res_dict = {}
                    for item in batch_mapped:
                        p_id = str(item.get("專利公開公告號", "")).strip()
                        res_dict[p_id] = item
                    
                    results = []
                    for idx, p_no, row in batch:
                        item = res_dict.get(p_no)
                        if item and item.get("AI技術簡述"):
                            results.append((p_no, {
                                "AI技術簡述": item.get("AI技術簡述", ""),
                                "技術特徵手段": item.get("技術特徵手段", ""),
                                "解決的技術問題或技術效益": item.get("解決的技術問題或技術效益", ""),
                                "初篩結果": "Y" if not enable_screening else ("Y" if str(item.get("初篩結果", "N")).upper() == "Y" else "N")
                            }))
                        else:
                            # 專利在批次回傳中遺漏或空白，啟動單件重試（最多 5 次）
                            logger.warning(f"專利 {p_no} 在批次回傳中遺漏或解析失敗，啟動單件重試機制...")
                            retry_result = await preprocess_single_patent_with_retry(
                                p_no, row,
                                title_col, abstract_col, novelty_col, use_col, advantage_col, claim_col,
                                enable_screening, screening_criteria, provider, client
                            )
                            results.append((p_no, retry_result))
                    return results
                except Exception as e:
                    logger.error(f"Preprocess batch failed: {e}. Attempting per-patent individual retry...")
                    results = []
                    for idx, p_no, row in batch:
                        # 整個批次 API 呼叫失敗，對每件專利單件重試
                        retry_result = await preprocess_single_patent_with_retry(
                            p_no, row,
                            title_col, abstract_col, novelty_col, use_col, advantage_col, claim_col,
                            enable_screening, screening_criteria, provider, client
                        )
                        results.append((p_no, retry_result))
                    return results

        # 預先建立 skipped 專利號的 set，供進度計算使用
        skipped_pnos = {r[1] for r in skipped_rows}

        # 執行 AI 全量預處理批次（ai_full_rows）
        if pending_rows:
            batches = [pending_rows[k:k+batch_size] for k in range(0, len(pending_rows), batch_size)]
            for batch in batches:
                batch_res = await process_batch(batch)
                for p_no, res in batch_res:
                    completed_patents[p_no] = res
                
                checkpoint_data["completed_patents"] = completed_patents
                with open(checkpoint_path, "w", encoding="utf-8") as f:
                    json.dump(checkpoint_data, f, ensure_ascii=False, indent=2)

                # 進度：已完成全量AI數量（不含skipped）
                ai_done = sum(1 for pno in completed_patents if pno not in skipped_pnos)
                task_registry[task_id]["completed_count"] = ai_done

        # 執行「純初篩」批次（ai_screen_only_rows）
        # 這批專利3欄均有值，僅需AI判斷Y/N初篩結果
        if ai_screen_only_rows:
            logger.info(f"開始對 {len(ai_screen_only_rows)} 件3欄有值但初篩缺漏的專利進行純初篩判斷...")
            for idx, p_no, row in ai_screen_only_rows:
                brief, means, effect = _get_existing_fields(row)
                title = str(row.get(title_col, "")).strip() if title_col else "無"
                screen_prompt = f"""你是專利分析與檢索專家。請針對以下 1 件專利，執行【落入範圍初篩作業】：
判定該專利是否落入特定的篩選準則範圍。初篩判定結果限為落入："Y" 或不落入："N"。

【篩選準則】：
{screening_criteria}

專利公開公告號: {p_no}
標題: {title}
AI技術簡述: {brief}
技術特徵手段: {means}
解決的技術問題或技術效益: {effect}

請輸出 JSON 物件，包含欄位：
  "專利公開公告號": "{p_no}"
  "AI技術簡述": "{brief}"
  "技術特徵手段": "{means}"
  "解決的技術問題或技術效益": "{effect}"
  "初篩結果": "Y" 或 "N"
"""
                try:
                    response_text = await safe_query_gemini_with_backoff(
                        screen_prompt, provider, client, response_schema=PatentPreprocessItem
                    )
                    parsed = robust_json_decode(response_text)
                    screen_val = "Y" if str(parsed.get("初篩結果", "N")).upper() == "Y" else "N"
                except Exception as e:
                    logger.warning(f"純初篩判斷失敗 ({p_no}): {e}，預設為 N")
                    screen_val = "N"

                completed_patents[p_no] = {
                    "AI技術簡述": brief,
                    "技術特徵手段": means,
                    "解決的技術問題或技術效益": effect,
                    "初篩結果": screen_val
                }

                # 更新進度
                ai_done = sum(1 for pno_key in completed_patents if pno_key not in skipped_pnos)
                task_registry[task_id]["completed_count"] = ai_done

            checkpoint_data["completed_patents"] = completed_patents
            with open(checkpoint_path, "w", encoding="utf-8") as f:
                json.dump(checkpoint_data, f, ensure_ascii=False, indent=2)

        for col in ["AI技術簡述", "技術特徵手段", "解決的技術問題或技術效益", "初篩結果"]:
            if col not in df.columns:
                df[col] = ""

        y_count = 0
        n_count = 0
        patents_preview = []
        for idx, row in df.iterrows():
            p_no = str(row.get(pub_col_name, "")).strip()
            if not p_no:
                p_no = f"ROW_{idx}"
            info = completed_patents.get(p_no, {
                "AI技術簡述": "",
                "技術特徵手段": "",
                "解決的技術問題或技術效益": "",
                "初篩結果": "Y"
            })
            df.at[idx, "AI技術簡述"] = info["AI技術簡述"]
            df.at[idx, "技術特徵手段"] = info["技術特徵手段"]
            df.at[idx, "解決的技術問題或技術效益"] = info["解決的技術問題或技術效益"]
            df.at[idx, "初篩結果"] = info["初篩結果"]
            
            if info["初篩結果"] == "Y":
                y_count += 1
            else:
                n_count += 1
            
            patents_preview.append({
                "專利公開公告號": p_no,
                "標題": str(row.get(title_col, "")) if title_col else "",
                "AI技術簡述": info["AI技術簡述"],
                "技術特徵手段": info["技術特徵手段"],
                "解決的技術問題或技術效益": info["解決的技術問題或技術效益"],
                "初篩結果": info["初篩結果"]
            })

        out_filename = f"preprocessed_{filename}"
        out_dir = os.path.join(tempfile.gettempdir(), "mindmap_preprocess", file_id)
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, out_filename)
        
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False)
            df_criteria = pd.DataFrame([
                {"篩選設定項目": "是否啟用初篩", "設定內容": "是" if enable_screening else "否"},
                {"篩選設定項目": "篩選準則內容", "設定內容": screening_criteria if screening_criteria else "無"}
            ])
            df_criteria.to_excel(writer, sheet_name="初篩定義", index=False)

        old_session_id = temp_storage.get(file_id, {}).get("x_session_id", "")
        temp_storage[file_id] = {
            "df_preprocessed": df,
            "file_path_preprocessed": out_path,
            "filename": filename,
            "x_session_id": old_session_id
        }

        hit_rate = (y_count / total_count) * 100 if total_count > 0 else 0.0

        task_registry[task_id]["status"] = "completed"
        task_registry[task_id]["y_count"] = y_count
        task_registry[task_id]["n_count"] = n_count
        task_registry[task_id]["hit_rate"] = f"{hit_rate:.1f}%"
        task_registry[task_id]["result"] = {
            "file_id": file_id,
            "filename": filename,
            "y_count": y_count,
            "n_count": n_count,
            "hit_rate": f"{hit_rate:.1f}%",
            "patents": patents_preview
        }

        # Log patent count to session if available
        session_id = temp_storage.get(file_id, {}).get("x_session_id", "")
        if session_id:
            try:
                from logger_handler import increment_patents_processed
                await increment_patents_processed(session_id, total_count)
            except Exception as e:
                logger.warning(f"Failed to log patent count to session: {e}")

        if os.path.exists(checkpoint_path):
            try:
                os.remove(checkpoint_path)
            except:
                pass

    except Exception as err:
        logger.error(f"Preprocess background task failed: {err}")
        task_registry[task_id]["status"] = "failed"
        task_registry[task_id]["error"] = str(err)


def detect_and_parse_bypass_excel(file_path, filename, file_id):
    """
    偵測上傳的 Excel 是否為先前導出的心智圖結果。
    如果是，則解析內容並重構最終的分類心智圖結構，回傳 (True, final_result)。
    否則回傳 (False, None)。
    """
    try:
        xls = pd.ExcelFile(file_path)
        sheet_name = "專利映射結果" if "專利映射結果" in xls.sheet_names else xls.sheet_names[0]
        df = xls.parse(sheet_name).fillna("")
        
        required_cols = [
            "專利公開公告號", "技術1階", "技術2階", "技術3階", 
            "應用領域", "功效節點", "AI技術簡述", "技術特徵手段", 
            "解決的技術問題或技術效益"
        ]
        
        # 欄位比對：忽略前後空白，但字元需完全一致
        col_mapping = {}
        for r_col in required_cols:
            found = next((c for c in df.columns if str(c).strip() == r_col), None)
            if found is not None:
                col_mapping[r_col] = found
        
        if len(col_mapping) < len(required_cols):
            return False, None
            
        import ast
        def parse_list(val):
            if not val:
                return []
            val_str = str(val).strip()
            if (val_str.startswith('[') and val_str.endswith(']')) or (val_str.startswith('(') and val_str.endswith(')')):
                try:
                    parsed = ast.literal_eval(val_str)
                    if isinstance(parsed, (list, tuple)):
                        return [str(x).strip() for x in parsed]
                    elif isinstance(parsed, set):
                        return [str(x).strip() for x in parsed]
                except Exception:
                    pass
            for sep in (',', ';', '、'):
                if sep in val_str:
                    return [x.strip() for x in val_str.split(sep)]
            return [val_str]

        patents = []
        for idx, row in df.iterrows():
            p_no = str(row.get(col_mapping["專利公開公告號"], "")).strip()
            if not p_no:
                p_no = f"ROW_{idx}"
            
            t1 = parse_list(row.get(col_mapping["技術1階"]))
            t2 = parse_list(row.get(col_mapping["技術2階"]))
            t3 = parse_list(row.get(col_mapping["技術3階"]))
            app = parse_list(row.get(col_mapping["應用領域"]))
            eff = parse_list(row.get(col_mapping["功效節點"]))
            
            patents.append({
                "專利公開公告號": p_no,
                "技術1階": t1,
                "技術2階": t2,
                "技術3階": t3,
                "應用領域": app,
                "功效節點": eff,
                "AI技術簡述": str(row.get(col_mapping["AI技術簡述"], "")).strip(),
                "技術特徵手段": str(row.get(col_mapping["技術特徵手段"], "")).strip(),
                "解決的技術問題或技術效益": str(row.get(col_mapping["解決的技術問題或技術效益"], "")).strip()
            })

        definitions = {}
        if "分類標籤定義" in xls.sheet_names:
            df_defs = xls.parse("分類標籤定義").fillna("")
            name_col = next((c for c in df_defs.columns if str(c).strip() == "分類名稱"), None)
            def_col = next((c for c in df_defs.columns if str(c).strip() == "定義說明"), None)
            if name_col and def_col:
                for _, row in df_defs.iterrows():
                    name = str(row.get(name_col, "")).strip()
                    definition = str(row.get(def_col, "")).strip()
                    if name:
                        definitions[name] = definition

        # 重構技術樹層階 (1-2階)
        t1_to_t2 = {}
        for p in patents:
            t1_val = p["技術1階"][0] if p["技術1階"] else "其他"
            t2_val = p["技術2階"][0] if p["技術2階"] else "其他"
            t1_to_t2.setdefault(t1_val, set()).add(t2_val)
            
        taxonomy_tech_tree = []
        for t1, t2_set in t1_to_t2.items():
            taxonomy_tech_tree.append({
                "技術1階": t1,
                "技術2階": sorted(list(t2_set))
            })

        all_apps = set()
        all_effs = set()
        for p in patents:
            for a in p["應用領域"]:
                if a: all_apps.add(a)
            for e in p["功效節點"]:
                if e: all_effs.add(e)

        # 優先從「專利映射結果」工作表的 N1 儲存格讀取 summary_title
        summary_title = ""
        try:
            from openpyxl import load_workbook
            wb_read = load_workbook(file_path, read_only=True, data_only=True)
            if "專利映射結果" in wb_read.sheetnames:
                ws_read = wb_read["專利映射結果"]
                n1_val = ws_read["N1"].value
                if n1_val and str(n1_val).strip():
                    summary_title = str(n1_val).strip()
            wb_read.close()
        except Exception as e:
            logger.warning(f"Failed to read summary_title from N1: {e}")

        # 若 N1 無值，則從檔名推導
        if not summary_title:
            summary_title, _ = os.path.splitext(filename)
            if not summary_title or summary_title.lower().startswith("preprocessed_") or summary_title.lower().startswith("mind_map_"):
                summary_title = "專利分類心智圖"

        stage1_taxonomy = {
            "summary_title": summary_title,
            "應用領域": sorted(list(all_apps)),
            "功效節點": sorted(list(all_effs)),
            "技術樹": taxonomy_tech_tree,
            "定義說明": definitions
        }

        final_result = {
            "summary_title": summary_title,
            "patents": patents,
            "file_id": file_id,
            "filename": filename,
            "stage1_taxonomy": stage1_taxonomy
        }

        return True, final_result
    except Exception as e:
        logger.error(f"Error parsing bypass excel: {e}")
        return False, None


@router.post("/api/mindmap/preprocess")
async def preprocess_patent_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    enable_screening: bool = Form(False),
    screening_criteria: str = Form(""),
    x_session_id: str = Form(default="")
):
    try:
        file_id = str(uuid.uuid4())
        temp_dir = os.path.join(tempfile.gettempdir(), "mindmap_preprocess", file_id)
        os.makedirs(temp_dir, exist_ok=True)
        file_path = os.path.join(temp_dir, file.filename)
        with open(file_path, "wb") as buffer:
            buffer.write(await file.read())

        # Log uploaded filename to session
        if x_session_id:
            try:
                from logger_handler import add_uploaded_file
                await add_uploaded_file(x_session_id, file.filename)
            except Exception:
                pass
            
        # 1. 偵測是否可直接 Bypass AI 分析流程
        is_bypass, final_result = detect_and_parse_bypass_excel(file_path, file.filename, file_id)
        if is_bypass:
            logger.info("偵測到已分析之 Excel 檔案，啟動直接生成分類心智圖 (Bypass) 流程。")
            df_patents = pd.DataFrame(final_result["patents"])
            temp_storage[file_id] = {
                "df": df_patents,
                "text": df_to_ai_content(df_patents),
                "filename": file.filename,
                "file_path_preprocessed": file_path,
                "stage1_taxonomy": final_result["stage1_taxonomy"],
                "stage2_result": final_result,
                "x_session_id": x_session_id
            }
            task_id = str(uuid.uuid4())
            task_registry[task_id] = {
                "status": "completed",
                "completed_count": len(final_result["patents"]),
                "total_count": len(final_result["patents"]),
                "result": {
                    "is_bypass": True,
                    "tree_data": final_result
                }
            }
            # Log patent count in bypass mode
            if x_session_id:
                try:
                    from logger_handler import increment_patents_processed
                    await increment_patents_processed(x_session_id, len(final_result["patents"]))
                except Exception:
                    pass
            return {"task_id": task_id, "file_id": file_id, "status": "completed"}

        # 2. 正常預處理流程
        try:
            df = extract_patents_from_excel(file_path)
        except Exception as e:
            logger.error(f"Failed to read excel file: {e}")
            raise HTTPException(status_code=400, detail="上傳檔案無相關資料，請重新上傳檔案")

        if df.empty:
            raise HTTPException(status_code=400, detail="上傳檔案無相關資料，請重新上傳檔案")

        # 欄位校驗：若同時沒有 1. 專利公開公告號 2. 專利摘要 3. First Claim 欄位，則報錯
        pub_col_exists = any("publication number" in str(c).lower() or "號" in str(c).lower() or "number" in str(c).lower() for c in df.columns)
        abstract_col_exists = any(any(k in str(c).lower() for k in ["摘要", "abstract"]) for c in df.columns)
        claim_col_exists = any(any(k in str(c).lower() for k in ["claim", "申請專利範圍", "權利要求", "first claim"]) for c in df.columns)

        if not pub_col_exists and not abstract_col_exists and not claim_col_exists:
            raise HTTPException(status_code=400, detail="上傳檔案無相關資料，請重新上傳檔案")

        task_id = str(uuid.uuid4())
        task_registry[task_id] = {
            "status": "processing",
            "completed_count": 0,
            "total_count": len(df),
            "result": None
        }

        # Store session_id for background task to log patent count
        temp_storage[file_id] = {"x_session_id": x_session_id}

        background_tasks.add_task(
            run_preprocess_task,
            task_id,
            file_id,
            file_path,
            file.filename,
            enable_screening,
            screening_criteria
        )

        return {"task_id": task_id, "file_id": file_id, "status": "processing"}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Preprocess startup failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/mindmap/preprocess_status")
async def get_preprocess_status(task_id: str):
    if task_id not in task_registry:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task_info = task_registry[task_id]
    return {
        "status": task_info["status"],
        "completed_count": task_info["completed_count"],
        "total_count": task_info["total_count"],
        "y_count": task_info.get("y_count", 0),
        "n_count": task_info.get("n_count", 0),
        "hit_rate": task_info.get("hit_rate", "0.0%"),
        "error": task_info.get("error"),
        "result": task_info.get("result")
    }


@router.post("/api/mindmap/start_from_preprocess")
async def start_from_preprocess(req: StartFromPreprocessRequest):
    try:
        file_id = req.file_id
        if file_id not in temp_storage:
            raise HTTPException(status_code=400, detail="Session expired or invalid file_id")
        
        session = temp_storage[file_id]
        df_full = session["df_preprocessed"]
        
        df_filtered = df_full[df_full["初篩結果"] == "Y"].copy()
        
        if df_filtered.empty:
            raise HTTPException(status_code=400, detail="初篩結果落入(Y)的專利數量為0，無法生成心智圖分類。")

        text_content = df_to_ai_content(df_filtered)
        
        result = query_gemini_stage1(text_content, req.config, df=df_filtered)
        if not result:
            raise ValueError("AI processing failed.")
            
        result["patents"] = []
        
        session["df"] = df_filtered
        session["text"] = text_content
        session["stage1_result"] = result
        
        result["file_id"] = file_id
        result["filename"] = session["filename"]
        result["is_stage1"] = True
        return result
    except Exception as e:
        logger.error(f"Failed to start classification from preprocessed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/mindmap/export_preprocessed")
async def export_preprocessed(file_id: str, x_session_id: str = Header(default="")):
    if file_id not in temp_storage or "file_path_preprocessed" not in temp_storage[file_id]:
        raise HTTPException(status_code=400, detail="Preprocessed file not found or expired.")
    
    file_path = temp_storage[file_id]["file_path_preprocessed"]
    filename = f"preprocessed_{temp_storage[file_id]['filename']}"
    


    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/api/mindmap/ai_assist_criteria")
async def ai_assist_criteria(req: AIAssistCriteriaRequest):
    try:
        import dotenv
        dotenv.load_dotenv(override=True)
        provider = os.environ.get("API_PROVIDER", "gemini").lower().strip().replace('"', '').replace("'", "")
        client = get_genai_client() if provider != "openrouter" else None
        
        type_names = {
            "fishbone": "技術魚骨1-2階節點",
            "di_query": "DI檢索式 (Derwent Innovation)",
            "keywords": "技術Keywords"
        }
        type_name = type_names.get(req.criteria_type, req.criteria_type)
        
        prompt = f"""你是專利分析與檢索專家。
請依據使用者選擇的篩選準則類型「{type_name}」以及所輸入的內容，生成一段約 200 字左右的繁體中文「專利技術篩選準則範圍描述」。

【篩選準則類型】：{type_name}
【使用者輸入的內容】：{req.user_input}

【技術關係分析規範】：
- 請仔細解析使用者輸入的關鍵字或內容。
- 如果輸入的關鍵字中含有逗號、分號、頓號、大於符號、斜線或上下層級縮進等，請合理分析並識別出它們之間的「上下階層關係（如父子、主從關係）」或「平行階層關係（如並列、替代關係）」。
- 例如：若為技術魚骨節點，分析其一階與二階技術分支的層級關係；若為檢索式或關鍵字，分析其邏輯組合關係（AND/OR）。

【生成要求】：
1. 請產出一篇通順、語意連貫的繁體中文段落（長度約 200 字左右）。
2. 段落中必須包含使用者輸入的所有核心 Keyword，並簡要描述這些技術點之間的對應關係與互動邏輯。
3. 輸出必須**僅包含**生成的技術範圍描述段落本身，嚴禁任何引言、結語或 Markdown 區塊。
"""
        generated_text = await safe_query_gemini_with_backoff(prompt, provider, client)
        if not generated_text:
            raise HTTPException(status_code=500, detail="AI 生成失敗")
            
        return {"generated_text": generated_text.strip()}
    except Exception as e:
        logger.error(f"AI assist criteria failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

temp_storage = {}

@router.post("/api/mindmap/upload")
async def upload_for_mindmap(file: UploadFile = File(...)):
    try:
        file_id = str(uuid.uuid4())
        temp_dir = os.path.join(tempfile.gettempdir(), "mindmap", file_id)
        os.makedirs(temp_dir, exist_ok=True)
        file_path = os.path.join(temp_dir, file.filename)
        with open(file_path, "wb") as buffer:
            buffer.write(await file.read())
        try:
            df = extract_patents_from_excel(file_path)
        except Exception as e:
            logger.error(f"Failed to read excel file: {e}")
            raise HTTPException(status_code=400, detail="上傳檔案無相關資料，請重新上傳檔案")

        if df.empty:
            raise HTTPException(status_code=400, detail="上傳檔案無相關資料，請重新上傳檔案")

        # 欄位校驗：若同時沒有 1. 專利公開公告號 2. 專利摘要 3. First Claim 欄位，則報錯
        pub_col_exists = any("publication number" in str(c).lower() or "號" in str(c).lower() or "number" in str(c).lower() for c in df.columns)
        abstract_col_exists = any(any(k in str(c).lower() for k in ["摘要", "abstract"]) for c in df.columns)
        claim_col_exists = any(any(k in str(c).lower() for k in ["claim", "申請專利範圍", "權利要求", "first claim"]) for c in df.columns)

        if not pub_col_exists and not abstract_col_exists and not claim_col_exists:
            raise HTTPException(status_code=400, detail="上傳檔案無相關資料，請重新上傳檔案")

        text_content = df_to_ai_content(df)
        
        config = MindMapConfig()
        result = query_gemini_stage1(text_content, config, df=df)
        if not result: raise ValueError("AI processing failed.")
        
        # Clean and validate Stage 1 results (no patents mapping at this stage)
        result["patents"] = []

        temp_storage[file_id] = {
            "text": text_content,
            "df": df,
            "filename": file.filename,
            "stage1_result": result
        }
        
        result["file_id"] = file_id
        result["filename"] = file.filename
        result["is_stage1"] = True
        return result
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/mindmap/reprocess")
async def reprocess_mindmap(config: MindMapConfig):
    try:
        if config.file_id not in temp_storage: raise ValueError("Session expired.")
        text_content = temp_storage[config.file_id]["text"]
        df = temp_storage[config.file_id].get("df")
        result = query_gemini_stage1(text_content, config, df=df)
        if not result: raise ValueError("AI processing failed.")
        
        result["patents"] = []

        temp_storage[config.file_id]["stage1_result"] = result
        
        result["file_id"] = config.file_id
        result["filename"] = temp_storage[config.file_id]["filename"]
        result["is_stage1"] = True
        return result
    except Exception as e:
        logger.error(f"Reprocess error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class MapStage1Request(BaseModel):
    file_id: str
    taxonomy: dict
    resume: bool = True  # True = 接續上次進度；False = 刪除舊 checkpoint 重新執行

@router.get("/api/mindmap/check_checkpoint")
async def check_checkpoint(file_id: str):
    """
    查詢指定 file_id 是否存在未完成的 checkpoint。
    回傳 has_checkpoint、已完成的 stage1/stage2 筆數，供前端決定是否顯示接續提示。
    """
    cp_path = os.path.join(checkpoint_dir, f"{file_id}_checkpoint.json")
    if not os.path.exists(cp_path):
        return {"has_checkpoint": False}
    try:
        with open(cp_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        s1_count = len(data.get("stage1_completed", []))
        s2_count = len(data.get("stage2_completed", []))
        total = s1_count  # 以 stage1 已完成數作為參考基準
        return {
            "has_checkpoint": True,
            "stage1_completed_count": s1_count,
            "stage2_completed_count": s2_count,
        }
    except Exception as e:
        logger.warning(f"check_checkpoint: failed to read checkpoint for {file_id}: {e}")
        return {"has_checkpoint": False}

@router.post("/api/mindmap/map_stage1")
async def map_stage1(req: MapStage1Request, background_tasks: BackgroundTasks):
    try:
        file_id = req.file_id
        if file_id not in temp_storage:
            raise HTTPException(status_code=400, detail="Session expired or invalid file_id")
        
        session = temp_storage[file_id]
        df = session["df"]

        # 若使用者選擇「重新執行」，先刪除舊的 checkpoint
        if not req.resume:
            cp_path = os.path.join(checkpoint_dir, f"{file_id}_checkpoint.json")
            if os.path.exists(cp_path):
                try:
                    os.remove(cp_path)
                    logger.info(f"Checkpoint for file_id={file_id} deleted (resume=False).")
                except Exception as e:
                    logger.warning(f"Failed to delete checkpoint for {file_id}: {e}")
        
        task_id = str(uuid.uuid4())
        task_registry[task_id] = {
            "status": "processing",
            "completed_count": 0,
            "total_count": len(df) * 2,
            "result": None
        }
        
        config = MindMapConfig()
        
        background_tasks.add_task(
            run_full_mindmap_task,
            task_id,
            df,
            req.taxonomy,
            file_id,
            config
        )
        
        return {"task_id": task_id, "status": "processing"}
    except Exception as e:
        logger.error(f"Failed to start stage 1 mapping task: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/mindmap/task_status")
async def get_task_status(task_id: str):
    if task_id not in task_registry:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task_info = task_registry[task_id]
    return {
        "status": task_info["status"],
        "stage": task_info.get("stage", 1),
        "completed_count": task_info["completed_count"],
        "total_count": task_info["total_count"],
        "error": task_info.get("error")
    }

def execute_batch_mapping(df, taxonomy, provider, client=None):
    # 優先抓取 "Publication Number" 欄位，其次抓含「號」的欄位
    pub_col_name = (
        next((c for c in df.columns if "publication number" in str(c).lower()), None)
        or next((c for c in df.columns if "號" in str(c)), None)
    )
    col_app = next((c for c in df.columns if "應用領域" in str(c)), "")
    col_eff = next((c for c in df.columns if "功效節點" in str(c)), "")
    col_brief = next((c for c in df.columns if "AI技術簡述" in str(c)), "")
    col_means = next((c for c in df.columns if "技術特徵手段" in str(c)), "")
    col_effect = next((c for c in df.columns if "解決的技術問題或技術效益" in str(c)), "")

    patents_data = []
    for _, row in df.iterrows():
        p_no = str(row.get(pub_col_name, "N/A")).strip()
        bits = []
        if col_app and str(row[col_app]).strip(): bits.append(f"原領域:{str(row[col_app]).strip()}")
        if col_eff and str(row[col_eff]).strip(): bits.append(f"原功效:{str(row[col_eff]).strip()}")
        
        tech_elements = []
        if col_brief and str(row[col_brief]).strip(): tech_elements.append(f"簡述:{str(row[col_brief]).strip()}")
        if col_means and str(row[col_means]).strip(): tech_elements.append(f"手段:{str(row[col_means]).strip()}")
        if col_effect and str(row[col_effect]).strip(): tech_elements.append(f"效益:{str(row[col_effect]).strip()}")
        if tech_elements:
            bits.append(f"技術特徵:{' '.join(tech_elements)}")
            
        patents_data.append((p_no, " | ".join(bits)))

    mapped_patents = []
    batch_size = 50
    for i in range(0, len(patents_data), batch_size):
        batch = patents_data[i:i+batch_size]
        batch_text = ""
        for p_no, details in batch:
            batch_text += f"[{p_no}] {details}\n"

        prompt = f"""你是專利分類專家。請將這批專利對應到以下已定義的分類目錄中。
請嚴格遵守：專利的「技術路徑」中的技術1階與技術2階名稱、「應用領域」名稱、「功效節點」名稱，必須完全來自下方目錄，不可自行拼寫、創造或擴充。

【分類目錄】：
1. 應用領域：{json.dumps(taxonomy.get("應用領域", []), ensure_ascii=False)}
2. 功效節點：{json.dumps(taxonomy.get("功效節點", []), ensure_ascii=False)}
3. 技術樹：
{json.dumps(taxonomy.get("技術樹", []), ensure_ascii=False)}

【任務要求】：
- 為每件專利指派一個或多個「應用領域」（必須是目錄中存在的）。
- 為每件專利指派一個或多個「功效節點」（必須是目錄中存在的）。
- 為每件專利指派一條或多條「技術路徑」，技術路徑必須是由目錄中存在的技術1階與其所屬的技術2階組成的完整陣列，如：["技術1階名稱", "技術2階名稱"]。

【格式】ONLY output a JSON object with a single key "patents" (array of objects).
例如：
{{
  "patents": [
    {{
      "專利公開公告號": "...",
      "技術路徑": [
        ["技術1階名稱", "技術2階名稱"]
      ],
      "應用領域": ["領域1"],
      "功效節點": ["功效A"]
    }}
  ]
}}

待對應專利列表：
{batch_text}
"""
        try:
            config_params = configure_response_format(Stage1MappingResponse, provider)
            response_text = ""
            if provider == "openrouter":
                payload = {
                    "model": "google/gemini-2.5-flash",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1, "max_tokens": 8000,
                    **config_params
                }
                resp_data = analyzer.send_openrouter_request(payload, timeout=300.0)
                response_text = resp_data["choices"][0]["message"]["content"]
            else:
                config_dict = {
                    "max_output_tokens": 8000,
                    "temperature": 0.1,
                    **config_params
                }
                resp = client.models.generate_content(
                    model="gemini-2.5-flash", contents=prompt,
                    config=config_dict
                )
                response_text = resp.text

            parsed = robust_json_decode(response_text)
            batch_mapped = parsed.get("patents", [])
            mapped_patents.extend(batch_mapped)
        except Exception as e:
            logger.error(f"Batch mapping failed for batch indices {i} to {i+batch_size}: {e}")
            for p_no, _ in batch:
                mapped_patents.append({
                    "專利公開公告號": p_no,
                    "技術路徑": [["其他", "其他"]],
                    "應用領域": ["其他"],
                    "功效節點": ["其他"]
                })
    return mapped_patents

checkpoint_dir = os.path.join(tempfile.gettempdir(), "mindmap_checkpoints")
os.makedirs(checkpoint_dir, exist_ok=True)

async def run_full_mindmap_task(task_id: str, df, taxonomy: dict, file_id: str, config: MindMapConfig):
    try:
        import dotenv
        dotenv.load_dotenv(override=True)
        provider = os.environ.get("API_PROVIDER", "gemini").lower().strip().replace('"', '').replace("'", "")
        client = get_genai_client() if provider != "openrouter" else None

        # 優先抓取 "Publication Number" 欄位，其次抓含「號」的欄位
        pub_col_name = (
            next((c for c in df.columns if "publication number" in str(c).lower()), None)
            or next((c for c in df.columns if "號" in str(c)), None)
        )
        col_app = next((c for c in df.columns if "應用領域" in str(c)), "")
        col_eff = next((c for c in df.columns if "功效節點" in str(c)), "")
        col_brief = next((c for c in df.columns if "AI技術簡述" in str(c)), "")
        col_means = next((c for c in df.columns if "技術特徵手段" in str(c)), "")
        col_effect = next((c for c in df.columns if "解決的技術問題或技術效益" in str(c)), "")

        patent_lookup = {}
        for _, row in df.iterrows():
            p_no = str(row.get(pub_col_name, "")).strip()
            patent_lookup[p_no] = {
                "AI技術簡述": str(row.get(col_brief, "")).strip(),
                "技術特徵手段": str(row.get(col_means, "")).strip(),
                "解決的技術問題或技術效益": str(row.get(col_effect, "")).strip(),
                "原領域": str(row.get(col_app, "")).strip(),
                "原功效": str(row.get(col_eff, "")).strip()
            }

        # Checkpoint 路徑使用 file_id 命名（而非 task_id），確保跨任務重啟後仍可找到
        checkpoint_path = os.path.join(checkpoint_dir, f"{file_id}_checkpoint.json")
        checkpoint_data = {
            "stage1_completed": [],
            "stage2_completed": [],
            "stage2_t3_categories": {}
        }

        # 嘗試從 Checkpoint 復原
        if os.path.exists(checkpoint_path):
            try:
                with open(checkpoint_path, "r", encoding="utf-8") as f:
                    loaded = json.load(f)
                    checkpoint_data["stage1_completed"] = loaded.get("stage1_completed", [])
                    checkpoint_data["stage2_completed"] = loaded.get("stage2_completed", [])
                    checkpoint_data["stage2_t3_categories"] = loaded.get("stage2_t3_categories", {})
                    logger.info(f"Task {task_id}: Checkpoint loaded successfully.")
            except Exception as e:
                logger.error(f"Task {task_id}: Failed to load checkpoint: {e}")

        # ----------------------------------------------------
        # 第一部分：Stage 1 專利映射 (每批 10 件)
        # ----------------------------------------------------
        stage1_mapped = checkpoint_data["stage1_completed"]
        stage1_pnos = {p["專利公開公告號"].strip() for p in stage1_mapped if p.get("專利公開公告號")}
        
        pending_stage1 = []
        for p_no, info in patent_lookup.items():
            if p_no not in stage1_pnos:
                bits = []
                if info["原領域"]: bits.append(f"原領域:{info['原領域']}")
                if info["原功效"]: bits.append(f"原功效:{info['原功效']}")
                
                tech_elements = []
                if info["AI技術簡述"]: tech_elements.append(f"簡述:{info['AI技術簡述']}")
                if info["技術特徵手段"]: tech_elements.append(f"手段:{info['技術特徵手段']}")
                if info["解決的技術問題或技術效益"]: tech_elements.append(f"效益:{info['解決的技術問題或技術效益']}")
                if tech_elements:
                    bits.append(f"技術特徵:{' '.join(tech_elements)}")
                
                pending_stage1.append((p_no, " | ".join(bits)))

        total_stage1 = len(patent_lookup)
        task_registry[task_id]["stage"] = 1
        task_registry[task_id]["total_count"] = total_stage1
        task_registry[task_id]["completed_count"] = len(stage1_mapped)

        if pending_stage1:
            batch_size = 10
            batches = [pending_stage1[k:k+batch_size] for k in range(0, len(pending_stage1), batch_size)]
            semaphore = asyncio.Semaphore(5)

            async def process_batch(batch):
                async with semaphore:
                    batch_text = ""
                    for p_no, details in batch:
                        batch_text += f"[{p_no}] {details}\n"

                    prompt = f"""你是專利分類專家。請將這批專利對應到以下已定義的分類目錄中。
請嚴格遵守：專利的「技術路徑」中的技術1階與技術2階名稱、「應用領域」名稱、「功效節點」名稱，必須完全來自下方目錄，不可自行拼寫、創造或擴充。

【分類目錄】：
1. 應用領域：{json.dumps(taxonomy.get("應用領域", []), ensure_ascii=False)}
2. 功效節點：{json.dumps(taxonomy.get("功效節點", []), ensure_ascii=False)}
3. 技術樹：
{json.dumps(taxonomy.get("技術樹", []), ensure_ascii=False)}

【任務要求】：
- 為每件專利指派一個或多個「應用領域」（必須是目錄中存在的）。
- 為每件專利指派一個或多個「功效節點」（必須是目錄中存在的）。
- 為每件專利指派一條或多條「技術路徑」，技術路徑必須是由目錄中存在的技術1階與其所屬的技術2階組成的完整陣列，如：["技術1階名稱", "技術2階名稱"]。

【格式】ONLY output a JSON object with a single key "patents" (array of objects).
例如：
{{
  "patents": [
    {{
      "專利公開公告號": "...",
      "技術路徑": [
        ["技術1階名稱", "技術2階名稱"]
      ],
      "應用領域": ["領域1"],
      "功效節點": ["功效A"]
    }}
  ]
}}

待對應專利列表：
{batch_text}
"""
                    try:
                        response_text = await safe_query_gemini_with_backoff(prompt, provider, client, response_schema=Stage1MappingResponse)
                        parsed = robust_json_decode(response_text)
                        batch_mapped = parsed.get("patents", [])
                        
                        # 標準化映射欄位
                        for p in batch_mapped:
                            if "專利公開公告號" not in p:
                                for k in ["公開公告號", "專利號", "patent_number"]:
                                    if k in p:
                                        p["專利公開公告號"] = p[k]
                                        break
                            p["專利公開公告號"] = str(p.get("專利公開公告號", "")).strip()
                            if "技術路徑" not in p or not p["技術路徑"]:
                                p["技術路徑"] = [["其他", "其他"]]
                            else:
                                new_paths = []
                                for path in p["技術路徑"]:
                                    if len(path) < 2:
                                        path = list(path) + ["其他"] * (2 - len(path))
                                    new_paths.append(path[:2])
                                p["技術路徑"] = new_paths
                            if "應用領域" not in p or not p["應用領域"]:
                                p["應用領域"] = ["其他"]
                            if "功效節點" not in p or not p["功效節點"]:
                                p["功效節點"] = ["其他"]
                        return batch_mapped
                    except Exception as e:
                        logger.error(f"Task {task_id}: Stage 1 batch mapping failed: {e}")
                        fallback_results = []
                        for p_no, _ in batch:
                            fallback_results.append({
                                "專利公開公告號": p_no,
                                "技術路徑": [["其他", "其他"]],
                                "應用領域": ["其他"],
                                "功效節點": ["其他"]
                            })
                        return fallback_results

            for idx, batch in enumerate(batches):
                batch_res = await process_batch(batch)
                stage1_mapped.extend(batch_res)
                checkpoint_data["stage1_completed"] = stage1_mapped
                
                # 更新進度
                task_registry[task_id]["completed_count"] = len(stage1_mapped)
                
                # 儲存 Checkpoint
                with open(checkpoint_path, "w", encoding="utf-8") as f:
                    json.dump(checkpoint_data, f, ensure_ascii=False, indent=2)

        # ----------------------------------------------------
        # 第二部分：Stage 2 技術 3 階生成與映射 (每組每批 10 件)
        # ----------------------------------------------------
        # 1. 依照 (T1, T2) 路徑分組專利號
        path_to_patents = {}
        for p in stage1_mapped:
            p_no = str(p.get("專利公開公告號", "")).strip()
            paths = p.get("技術路徑", [])
            for path in paths:
                if len(path) >= 2:
                    t1, t2 = path[0], path[1]
                    path_to_patents.setdefault((t1, t2), []).append(p_no)

        t3_categories = checkpoint_data["stage2_t3_categories"]
        stage2_mapped = checkpoint_data["stage2_completed"]

        # 建立已完成的 (專利公開公告號, 技術1階, 技術2階) 複合 Key 集合
        stage2_completed_keys = set()
        for p in stage2_mapped:
            p_no = p["專利公開公告號"].strip()
            t1 = p["技術1階"][0] if isinstance(p["技術1階"], list) and p["技術1階"] else p["技術1階"]
            t2 = p["技術2階"][0] if isinstance(p["技術2階"], list) and p["技術2階"] else p["技術2階"]
            stage2_completed_keys.add((p_no, t1, t2))

        # 計算 Stage 2 總「專利-路徑」任務筆數
        total_stage2_tasks = max(1, sum(len(p_nos) for p_nos in path_to_patents.values()))
        completed_stage2_tasks = len(stage2_completed_keys)

        task_registry[task_id]["stage"] = 2
        task_registry[task_id]["total_count"] = total_stage2_tasks
        task_registry[task_id]["completed_count"] = completed_stage2_tasks

        semaphore_s2 = asyncio.Semaphore(5)
        batch_size = 10

        for (t1, t2), p_nos in path_to_patents.items():
            if not p_nos:
                continue

            path_key = f"{t1} > {t2}"

            # a. 產生技術 3 階標籤 (如果不曾生成過)
            if path_key not in t3_categories:
                sample_pnos = p_nos[:40]
                batch_text_labels = ""
                for p_no in sample_pnos:
                    details = patent_lookup.get(p_no, {})
                    batch_text_labels += f"[{p_no}] 簡述:{details.get('AI技術簡述')} | 手段:{details.get('技術特徵手段')} | 效益:{details.get('解決的技術問題或技術效益')}\n"

                prompt_gen = f"""你是專利分類專家。現有以下專利被歸類於「技術1階：{t1}」->「技術2階：{t2}」下方。
請針對這批專利特徵，在「技術2階：{t2}」之下進一步細分出最匹配的「技術3階」子類別。

【分類限制與命名規則】：
1. 技術3階：在此「技術2階：{t2}」子樹下，最多只能細分出 {config.tech3_count} 個「技術3階」子類別。
2. 命名總字數限制：所有「技術3階」標籤之命名總字數不得超過 12 個字。
3. 避免重複上階名稱：下階（技術3階）之命名不須且不要再重複其上階（技術2階：{t2} 或技術1階：{t1}）已有的技術名稱。例如：若上階已包含「電路設計」，則技術3階不應命名為「電路設計模擬與驗證」，應簡化為「模擬與驗證」。

【格式】ONLY output a JSON object with a single key: "技術3階類別" (array of strings).
例如：
{{
  "技術3階類別": ["細分類別A", "細分類別B"]
}}

待分析專利清單：
{batch_text_labels}
"""
                try:
                    response_text_gen = await safe_query_gemini_with_backoff(prompt_gen, provider, client, response_schema=Stage2LabelResponse)
                    parsed_gen = robust_json_decode(response_text_gen)
                    t3_labels = parsed_gen.get("技術3階類別", [])
                    if not t3_labels:
                        t3_labels = ["其他"]
                    else:
                        cleaned_t3 = []
                        for t3 in t3_labels:
                            if t3:
                                t3_c = clean_label_name(t3, parent_label=t2)
                                t3_c = clean_label_name(t3_c, parent_label=t1)
                                if t3_c:
                                    cleaned_t3.append(t3_c)
                        t3_labels = cleaned_t3 if cleaned_t3 else ["其他"]
                    t3_categories[path_key] = t3_labels
                except Exception as e:
                    logger.error(f"Task {task_id}: Stage 2 Label Generation failed for {path_key}: {e}")
                    t3_categories[path_key] = ["其他"]

                checkpoint_data["stage2_t3_categories"] = t3_categories
                with open(checkpoint_path, "w", encoding="utf-8") as f:
                    json.dump(checkpoint_data, f, ensure_ascii=False, indent=2)

            t3_labels = t3_categories[path_key]

            # b. 分批映射 (每批 10 件)
            pending_s2_pnos = [p_no for p_no in p_nos if (p_no, t1, t2) not in stage2_completed_keys]
            if not pending_s2_pnos:
                continue

            batches_s2 = [pending_s2_pnos[k:k+batch_size] for k in range(0, len(pending_s2_pnos), batch_size)]

            async def process_s2_batch(batch_pnos):
                async with semaphore_s2:
                    batch_text_map = ""
                    for p_no in batch_pnos:
                        details = patent_lookup.get(p_no, {})
                        batch_text_map += f"[{p_no}] 簡述:{details.get('AI技術簡述')} | 手段:{details.get('技術特徵手段')} | 效益:{details.get('解決的技術問題或技術效益')}\n"

                    prompt_map = f"""你是專利分類專家。現有以下專利被歸類於「技術1階：{t1}」->「技術2階：{t2}」下方。
請將這批專利指派對應到指定的「技術3階」子類別中（必須來自下方給定的類別清單，不可自行發明）。

【技術3階子類別清單】：
{json.dumps(t3_labels, ensure_ascii=False)}

【格式】ONLY output a JSON object with a single key: "patents" (array of objects).
例如：
{{
  "patents": [
    {{
      "專利公開公告號": "...",
      "技術3階": ["細分類別A"]
    }}
  ]
}}

待對應專利清單：
{batch_text_map}
"""
                    try:
                        response_text_map = await safe_query_gemini_with_backoff(prompt_map, provider, client, response_schema=Stage2MappingResponse)
                        parsed_map = robust_json_decode(response_text_map)
                        patent_t3_list = parsed_map.get("patents", [])
                        
                        mapped_batch = []
                        for item in patent_t3_list:
                            p_id = str(item.get("專利公開公告號", "")).strip()
                            t3_vals = item.get("技術3階", [])
                            if isinstance(t3_vals, str):
                                t3_vals = [t3_vals]
                            valid_t3s = [v for v in t3_vals if v in t3_labels]
                            if not valid_t3s and t3_labels:
                                valid_t3s = [t3_labels[0]]
                            elif not valid_t3s:
                                valid_t3s = ["其他"]
                            
                            s1_p = next((x for x in stage1_mapped if x["專利公開公告號"] == p_id), {})
                            details = patent_lookup.get(p_id, {})
                            
                            for t3 in valid_t3s:
                                mapped_batch.append({
                                    "專利公開公告號": p_id,
                                    "技術1階": [t1],
                                    "技術2階": [t2],
                                    "技術3階": [t3],
                                    "應用領域": s1_p.get("應用領域", ["其他"]),
                                    "功效節點": s1_p.get("功效節點", ["其他"]),
                                    "AI技術簡述": details.get("AI技術簡述", ""),
                                    "技術特徵手段": details.get("技術特徵手段", ""),
                                    "解決的技術問題或技術效益": details.get("解決的技術問題或技術效益", "")
                                })
                        return mapped_batch
                    except Exception as e:
                        logger.error(f"Task {task_id}: Stage 2 batch mapping failed for batch {batch_pnos[0]}~: {e}")
                        fallback_batch = []
                        for p_id in batch_pnos:
                            s1_p = next((x for x in stage1_mapped if x["專利公開公告號"] == p_id), {})
                            details = patent_lookup.get(p_id, {})
                            fallback_batch.append({
                                "專利公開公告號": p_id,
                                "技術1階": [t1],
                                "技術2階": [t2],
                                "技術3階": ["其他"],
                                "應用領域": s1_p.get("應用領域", ["其他"]),
                                "功效節點": s1_p.get("功效節點", ["其他"]),
                                "AI技術簡述": details.get("AI技術簡述", ""),
                                "技術特徵手段": details.get("技術特徵手段", ""),
                                "解決的技術問題或技術效益": details.get("解決的技術問題或技術效益", "")
                            })
                        return fallback_batch

            for batch_pnos in batches_s2:
                batch_res = await process_s2_batch(batch_pnos)
                stage2_mapped.extend(batch_res)
                checkpoint_data["stage2_completed"] = stage2_mapped
                
                # 更新已完成的複合 Key 集合
                for item in batch_res:
                    p_id = item["專利公開公告號"].strip()
                    t1_val = item["技術1階"][0] if isinstance(item["技術1階"], list) and item["技術1階"] else item["技術1階"]
                    t2_val = item["技術2階"][0] if isinstance(item["技術2階"], list) and item["技術2階"] else item["技術2階"]
                    stage2_completed_keys.add((p_id, t1_val, t2_val))
                
                completed_stage2_tasks = len(stage2_completed_keys)
                
                # 更新進度
                task_registry[task_id]["completed_count"] = completed_stage2_tasks
                
                # 儲存 Checkpoint
                with open(checkpoint_path, "w", encoding="utf-8") as f:
                    json.dump(checkpoint_data, f, ensure_ascii=False, indent=2)

        # ----------------------------------------------------
        # 第三部分：任務完成與結果合併
        # ----------------------------------------------------
        s2_mapped_pnos = {p["專利公開公告號"].strip() for p in stage2_mapped}
        for p in stage1_mapped:
            p_no = p["專利公開公告號"].strip()
            if p_no not in s2_mapped_pnos:
                paths = p.get("技術路徑", [["其他", "其他"]])
                for path in paths:
                    t1, t2 = path[0], path[1]
                    details = patent_lookup.get(p_no, {})
                    stage2_mapped.append({
                        "專利公開公告號": p_no,
                        "技術1階": [t1],
                        "技術2階": [t2],
                        "技術3階": ["其他"],
                        "應用領域": p.get("應用領域", ["其他"]),
                        "功效節點": p.get("功效節點", ["其他"]),
                        "AI技術簡述": details.get("AI技術簡述", ""),
                        "技術特徵手段": details.get("技術特徵手段", ""),
                        "解決的技術問題或技術效益": details.get("解決的技術問題或技術效益", "")
                    })

        final_result = {
            "summary_title": taxonomy.get("summary_title", "專利分類心智圖"),
            "patents": stage2_mapped,
            "file_id": file_id,
            "filename": temp_storage[file_id]["filename"],
            "stage1_taxonomy": taxonomy
        }

        temp_storage[file_id]["stage1_taxonomy"] = taxonomy
        temp_storage[file_id]["stage1_patents"] = stage1_mapped
        temp_storage[file_id]["stage2_result"] = final_result

        task_registry[task_id]["status"] = "completed"
        task_registry[task_id]["stage"] = 2
        task_registry[task_id]["completed_count"] = total_stage2_tasks
        task_registry[task_id]["total_count"] = total_stage2_tasks
        task_registry[task_id]["result"] = final_result

        # 清除 Checkpoint
        if os.path.exists(checkpoint_path):
            try:
                os.remove(checkpoint_path)
            except:
                pass

    except Exception as err:
        logger.error(f"Background task {task_id} failed: {err}")
        task_registry[task_id]["status"] = "failed"
        task_registry[task_id]["error"] = str(err)

@router.post("/api/mindmap/generate_stage2")
async def generate_stage2(data: dict):
    try:
        file_id = data.get("file_id")
        if not file_id or file_id not in temp_storage:
            raise HTTPException(status_code=400, detail="Session expired or file_id invalid.")
            
        if "stage2_result" in temp_storage[file_id]:
            return temp_storage[file_id]["stage2_result"]
            
        raise HTTPException(status_code=400, detail="Stage 2 result is not generated yet. Please wait for the background mapping task to complete.")
    except Exception as e:
        logger.error(f"Stage 2 execution error: {e}")


@router.post("/api/mindmap/export")
async def export_mindmap_excel(data: dict, x_session_id: str = Header(default="")):
    patents = data.get("patents", [])
    original_filename = data.get("filename", "download")
    if not patents: raise HTTPException(status_code=400, detail="No data.")
    
    df_out = pd.DataFrame(patents)
    
    stage1_taxonomy = data.get("stage1_taxonomy", {})
    definitions = stage1_taxonomy.get("定義說明", {})
    
    def_rows = []
    # 應用領域
    for d in stage1_taxonomy.get("應用領域", []):
        def_rows.append({
            "分類類型": "應用領域",
            "分類名稱": d,
            "定義說明": definitions.get(d, "")
        })
    # 功效節點
    for e in stage1_taxonomy.get("功效節點", []):
        def_rows.append({
            "分類類型": "功效節點",
            "分類名稱": e,
            "定義說明": definitions.get(e, "")
        })
    # 技術樹 (1-2階)
    for t1_item in stage1_taxonomy.get("技術樹", []):
        t1_name = t1_item.get("技術1階", "")
        def_rows.append({
            "分類類型": "技術1階",
            "分類名稱": t1_name,
            "定義說明": definitions.get(t1_name, "")
        })
        for t2_name in t1_item.get("技術2階", []):
            def_rows.append({
                "分類類型": "技術2階",
                "分類名稱": t2_name,
                "定義說明": definitions.get(t2_name, "")
            })
            
    df_defs = pd.DataFrame(def_rows)
    
    name, _ = os.path.splitext(original_filename)
    out_filename = f"{name}_分類結果.xlsx"
    temp_dir = os.path.join(tempfile.gettempdir(), "mindmap_export")
    os.makedirs(temp_dir, exist_ok=True)
    out_path = os.path.join(temp_dir, out_filename)
    
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name="專利映射結果", index=False)
        if not df_defs.empty:
            df_defs.to_excel(writer, sheet_name="分類標籤定義", index=False)

    # 將 summary_title 寫入「專利映射結果」工作表的 N1 儲存格
    summary_title_val = data.get("summary_title") or stage1_taxonomy.get("summary_title", "")
    if summary_title_val:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(out_path)
            ws = wb["專利映射結果"]
            ws["N1"] = summary_title_val
            wb.save(out_path)
        except Exception as e:
            logger.warning(f"Failed to write summary_title to N1: {e}")



    return FileResponse(path=out_path, filename=out_filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

